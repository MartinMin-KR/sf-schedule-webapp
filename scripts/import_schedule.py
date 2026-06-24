from __future__ import annotations

import json
import re
import sys
import unicodedata
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


YEAR = 2026


def slugify(text: str) -> str:
    normalized = unicodedata.normalize("NFKD", text)
    ascii_text = normalized.encode("ascii", "ignore").decode("ascii").lower()
    ascii_text = re.sub(r"[^a-z0-9]+", "-", ascii_text).strip("-")
    return ascii_text


def unique_slug(base: str, used: set[str]) -> str:
    candidate = base
    index = 2
    while candidate in used:
        candidate = f"{base}-{index}"
        index += 1
    used.add(candidate)
    return candidate


def parse_place(raw_value: str) -> tuple[str, str | None]:
    cleaned = re.sub(r"\s+", " ", raw_value).strip()
    match = re.match(r"^(.*?)\s*\((.*?)\)\s*$", cleaned)
    if match:
      return match.group(1).strip(), match.group(2).strip()
    return cleaned, None


def resolve_slot(raw_header: str, occurrence: int) -> tuple[str, str, str]:
    date_match = re.search(r"(\d+)\.(\d+)", raw_header)
    month = int(date_match.group(1)) if date_match else 6
    day = int(date_match.group(2)) if date_match else 24
    date = f"{YEAR}-{month:02d}-{day:02d}"

    if "버스 교체" in raw_header:
        return date, "버스 교체", "13:10", "13:30"
    if "오전" in raw_header:
        return date, "오전 방문", "09:00", "11:20"
    if "점심" in raw_header or "중식" in raw_header:
        return date, "점심 일정", "12:00", "13:20"
    if "오후" in raw_header:
        return date, "오후 방문", "14:00", "16:40"
    if "석식" in raw_header:
        if occurrence == 1:
            return date, "저녁 일정", "18:00", "19:10"
        return date, "저녁 일정 2", "19:20", "20:20"
    if day == 28:
        return date, "귀국 이동", "08:00", "10:00"
    return date, "특별 일정", "18:00", "19:30"


def main() -> int:
    source = Path(sys.argv[1])
    project_root = Path(sys.argv[2])
    workbook = load_workbook(source, data_only=True)
    sheet = workbook["정규데이터"]
    rows = list(sheet.iter_rows(values_only=True))
    headers = list(rows[0])
    data_rows = rows[1:]

    group_counts: dict[str, int] = defaultdict(int)
    groups: list[dict[str, object]] = []
    members: list[dict[str, str]] = []
    companies: list[dict[str, str]] = []
    schedules: list[dict[str, object]] = []

    group_seen: set[str] = set()
    member_slugs: set[str] = set()
    company_slugs: set[str] = set()
    company_ids: dict[str, str] = {}
    header_occurrence: dict[str, int] = defaultdict(int)

    parsed_headers: list[tuple[int, str, str, str, str]] = []
    for column_index, raw_header in enumerate(headers[4:], start=4):
        header_occurrence[str(raw_header)] += 1
        date, slot_label, start_time, end_time = resolve_slot(
            str(raw_header), header_occurrence[str(raw_header)]
        )
        parsed_headers.append((column_index, date, slot_label, start_time, end_time))

    for row in data_rows:
        group_name = f"{row[2]}조"
        if group_name not in group_seen:
            groups.append(
                {
                    "id": f"group-{row[2]}",
                    "name": group_name,
                }
            )
            group_seen.add(group_name)
        group_counts[group_name] += 1

        member_base = slugify(str(row[3])) or f"member-{int(row[1]):03d}"
        member_id = unique_slug(member_base, member_slugs)
        members.append(
            {
                "id": member_id,
                "name": str(row[3]).strip(),
                "groupId": f"group-{row[2]}",
            }
        )

        for sequence, (column_index, date, slot_label, start_time, end_time) in enumerate(
            parsed_headers, start=1
        ):
            raw_value = row[column_index]
            if raw_value in (None, ""):
                continue

            place_name, note = parse_place(str(raw_value))
            if place_name not in company_ids:
                company_base = slugify(place_name) or f"company-{len(company_ids) + 1:02d}"
                company_id = unique_slug(company_base, company_slugs)
                company_ids[place_name] = company_id
                companies.append(
                    {
                        "id": company_id,
                        "name": place_name,
                        "description": "실제 버스 일정표에서 가져온 방문 장소입니다.",
                        "address": "",
                    }
                )

            item = {
                "personId": member_id,
                "date": date,
                "slotLabel": slot_label,
                "sequence": sequence,
                "companyId": company_ids[place_name],
                "startTime": start_time,
                "endTime": end_time,
            }
            if note:
                item["note"] = f"배정 차량: {note}"
            schedules.append(item)

    for group in groups:
        group["size"] = group_counts[group["name"]]

    data_dir = project_root / "src" / "data"
    data_dir.mkdir(parents=True, exist_ok=True)
    (data_dir / "groups.json").write_text(
        json.dumps(groups, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    (data_dir / "members.json").write_text(
        json.dumps(members, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    (data_dir / "companies.json").write_text(
        json.dumps(companies, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    (data_dir / "schedules.json").write_text(
        json.dumps(schedules, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )

    print(f"groups={len(groups)} members={len(members)} companies={len(companies)} schedules={len(schedules)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
